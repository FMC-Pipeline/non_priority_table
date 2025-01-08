import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from colored import format_excel_with_feeds

# Path file yang benar
file_path = os.path.join(os.path.dirname(__file__), "data.txt")

# Baca data dari file
with open(file_path, "r") as file:
    data = file.readlines()

# Membersihkan data dari baris kosong dan separator
lines = [line.strip() for line in data if line.strip() and not line.startswith("||||")]

# Ekstrak data ke dalam kolom
processed_data = []
for line in lines:
    parts = line.split("|")
    if len(parts) >= 5:
        event_date = parts[1]
        table_name = parts[0]
        start_date = parts[2]
        end_date = parts[3]
        value = parts[4]
        processed_data.append([table_name, event_date, start_date, end_date, value])

# Buat DataFrame dari data yang diproses
df = pd.DataFrame(processed_data, columns=["TABLE NAME", "EVENT DATE", "DATE TRANSACTION", "DATE AVAILABILITY", "NOW SIZE CONDITION"])

# Membuka atau membuat workbook
cwd = os.getcwd()
output_excel_path = cwd+'\hasil\output.xlsx' #implement dynamic directory (!IMPORTANT kalo belum ada directory hasil tolong di create dulu sebelum di run, belum di handle)
if os.path.exists(output_excel_path):
    wb = load_workbook(output_excel_path)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])  # Hapus sheet default

# Tambahkan sheet utama, Daily, Weekly, Monthly, dan Billing jika belum ada
if "Main" not in wb.sheetnames:
    main_ws = wb.create_sheet("Main")
else:
    main_ws = wb["Main"]

if "Daily" not in wb.sheetnames:
    daily_ws = wb.create_sheet("Daily")
else:
    daily_ws = wb["Daily"]

if "Weekly" not in wb.sheetnames:
    weekly_ws = wb.create_sheet("Weekly")
else:
    weekly_ws = wb["Weekly"]

if "Monthly" not in wb.sheetnames:
    monthly_ws = wb.create_sheet("Monthly")
else:
    monthly_ws = wb["Monthly"]

if "Billing" not in wb.sheetnames:
    billing_ws = wb.create_sheet("Billing")
else:
    billing_ws = wb["Billing"]

# Fungsi untuk menambahkan tabel ke sheet
def add_table_to_sheet(ws, table_name, group, start_row):
    # Tambahkan nama tabel
    ws.append([f"TABLE NAME: {table_name}"])

    # Tambahkan header kolom
    start_row += 1
    ws.append(["TABLE NAME", "DATE TRANSACTION", "DATE AVAILABILITY", "TIME AVAILABILITY", "NOW SIZE CONDITION"])

    # Tambahkan data
    start_row += 1
    for row in group.values.tolist():
        ws.append(row)
        start_row += 1

    # Tambahkan baris kosong setelah tabel
    ws.append([])
    return start_row

# Proses setiap tabel berdasarkan jumlah baris
main_row = 1
daily_row = 1
weekly_row = 1
monthly_row = 1
billing_row = 1

for table_name, group in df.groupby("TABLE NAME"):
    print(f"Processing table: {table_name} with {len(group)} rows.")  # Debugging info
    
    # Cek apakah nama tabel mengandung "bil" atau "billing"
    if "bil" in table_name.lower() or "billing" in table_name.lower():
        billing_row = add_table_to_sheet(billing_ws, table_name, group, billing_row)
    elif len(group) >= 10:
        daily_row = add_table_to_sheet(daily_ws, table_name, group, daily_row)
    elif 1 < len(group) < 6:
        weekly_row = add_table_to_sheet(weekly_ws, table_name, group, weekly_row)
    elif len(group) == 1:
        monthly_row = add_table_to_sheet(monthly_ws, table_name, group, monthly_row)
    else:
        main_row = add_table_to_sheet(main_ws, table_name, group, main_row)

# Menghapus teks "TABLE NAME: " dari kolom pertama di setiap sheet
for sheet in [main_ws, daily_ws, weekly_ws, monthly_ws, billing_ws]:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if row[0].value and "TABLE NAME:" in str(row[0].value):
            row[0].value = row[0].value.replace("TABLE NAME: ", "")

# Menyimpan workbook
wb.save(output_excel_path)
print(f"Data telah disimpan di {output_excel_path}")

# Format file Excel dengan coloured.py
print("Formatting Excel file with colours...")
output_colored_path = cwd + '\hasil\colored_output.xlsx'
format_excel_with_feeds(output_excel_path, output_colored_path)

# Konfirmasi bahwa file berwarna telah disimpan
print(f"Formatted Excel file saved at {output_colored_path}")

