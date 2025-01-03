import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_excel_with_feeds(file_path, output_path, autofit_columns=True):
    # Load workbook
    wb = openpyxl.load_workbook(file_path)

    # Define styles
    first_header_fill = PatternFill(start_color="3C7D22", end_color="3C7D22", fill_type="solid")
    second_header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Process each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        row = 1
        
        while row <= max_row:
            # Check if the row starts a feed block
            if sheet.cell(row=row, column=1).value and sheet.cell(row=row, column=1).value.strip():
                # Merge the feed name header across 5 columns
                feed_name = sheet.cell(row=row, column=1).value
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                
                cell = sheet.cell(row=row, column=1)
                cell.value = feed_name
                cell.fill = first_header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
                row += 1  # Move to the second header row

                # Style the second header row
                for col in range(1, 6):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = second_header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                row += 1  # Move to the data rows

                # Style the data rows until an empty row is encountered
                while row <= max_row and sheet.cell(row=row, column=1).value and sheet.cell(row=row, column=1).value.strip():
                    for col in range(1, 6):
                        cell = sheet.cell(row=row, column=col)
                        cell.border = thin_border
                    row += 1
            else:
                row += 1

        # Adjust column widths to fit the text if autofit is enabled
        if autofit_columns:
            for col in range(1, max_col + 1):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].auto_size = True

    # Save the formatted workbook
    wb.save(output_path)

# File paths (modify as needed)
input_file = "hasil/output.xlsx"
output_file = "hasil/colored_output.xlsx"

# Format the file
format_excel_with_feeds(input_file, output_file, autofit_columns=True)
